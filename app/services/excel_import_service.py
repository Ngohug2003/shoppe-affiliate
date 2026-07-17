from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from zipfile import BadZipFile

from openpyxl import load_workbook  # type: ignore[import-untyped]
from openpyxl.utils.exceptions import InvalidFileException  # type: ignore[import-untyped]

from app.utils.urls import extract_shopee_urls


class ExcelImportError(ValueError):
    pass


@dataclass(frozen=True)
class ExcelLinkExtraction:
    urls: list[str]
    scan_limit_reached: bool = False
    link_limit_reached: bool = False


class ExcelShopeeUrlExtractor:
    def __init__(self, *, max_links: int, max_cells: int) -> None:
        self.max_links = max_links
        self.max_cells = max_cells

    def extract(self, content: bytes) -> ExcelLinkExtraction:
        try:
            workbook = load_workbook(
                BytesIO(content),
                read_only=True,
                data_only=True,
            )
        except (BadZipFile, InvalidFileException, OSError, ValueError) as exc:
            raise ExcelImportError("File Excel không hợp lệ hoặc đã bị hỏng") from exc

        urls: list[str] = []
        seen: set[str] = set()
        scanned_cells = 0
        try:
            for worksheet in workbook.worksheets:
                for row in worksheet.iter_rows():
                    for cell in row:
                        scanned_cells += 1
                        if scanned_cells > self.max_cells:
                            return ExcelLinkExtraction(
                                urls=urls,
                                scan_limit_reached=True,
                            )
                        values = [cell.value]
                        hyperlink = getattr(cell, "hyperlink", None)
                        if hyperlink is not None:
                            values.append(getattr(hyperlink, "target", None))
                        for value in values:
                            if not isinstance(value, str):
                                continue
                            for url in extract_shopee_urls(value):
                                if url in seen:
                                    continue
                                if len(urls) >= self.max_links:
                                    return ExcelLinkExtraction(
                                        urls=urls,
                                        link_limit_reached=True,
                                    )
                                seen.add(url)
                                urls.append(url)
        finally:
            workbook.close()

        return ExcelLinkExtraction(urls=urls)
